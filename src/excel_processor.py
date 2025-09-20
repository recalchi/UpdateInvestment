# src/excel_processor.py
import os
import time
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

class ExcelProcessor:
    """
    ExcelProcessor responsável por:
    - detectar linha de header (0..4)
    - normalizar nomes de colunas (strip, upper, remover UNNAMED)
    - converter colunas numéricas conhecidas
    - formatar coluna de data (DATA ATT)
    - cache simples baseado em mtime + TTL para evitar leituras a cada request
    """
    def __init__(self, file_path: str, cache_ttl_seconds: int = 5):
        self.file_path = file_path
        self._cache_df = None
        self._cache_mtime = None
        self.cache_ttl_seconds = cache_ttl_seconds
        self._last_cache_time = 0

    def _file_mtime(self):
        try:
            return os.path.getmtime(self.file_path)
        except Exception:
            return None

    def _should_reload(self):
        mtime = self._file_mtime()
        if mtime is None:
            return True
        if self._cache_mtime is None or mtime != self._cache_mtime:
            return True
        if time.time() - self._last_cache_time > self.cache_ttl_seconds:
            return True
        return False

    def _normalize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        # remover colunas totalmente vazias
        df = df.dropna(how="all", axis=1)

        # normalizar nomes
        new_cols = []
        for c in df.columns:
            c_str = str(c).strip().upper().replace("\n", " ").replace("  ", " ")
            new_cols.append(c_str)
        df.columns = new_cols

        # remover 'UNNAMED' e colunas vazias
        cols_to_drop = [c for c in df.columns if c.startswith("UNNAMED") or c == ""]
        if cols_to_drop:
            df = df.drop(columns=cols_to_drop, errors="ignore")

        # mapear colunas comuns para nomes esperados
        mapping = {
            "ATIVO": "ATIVO",
            "TICKER": "TICKER",
            "PESO": "PESO",
            "DY": "DY",
            "PRECO_MAX": "ATÉ QUE PREÇO COMPRAR",  # <-- NOVO
            "ULTIMA_ATT": "DATA ATT",
            "RENTABILIDADE_ULT_MES": "RENTABILIDADE_ULT_MES",
            "RENTABILIDADE_ANUAL": "RENTABILIDADE_ANUAL"
        }
        rename_map = {c: mapping[c] for c in df.columns if c in mapping}
        if rename_map:
            df = df.rename(columns=rename_map)

        # converter colunas numéricas conhecidas
        for col in ["PESO", "DY", "RENTABILIDADE_ULT_MES", "RENTABILIDADE_ANUAL"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        # formatar DATA ATT: se existir, converter e preencher; senão criar com hoje
        if "DATA ATT" in df.columns:
            try:
                df["DATA ATT"] = pd.to_datetime(df["DATA ATT"], errors="coerce").dt.strftime("%Y-%m-%d")
                df["DATA ATT"] = df["DATA ATT"].fillna(pd.to_datetime("today").strftime("%Y-%m-%d"))
            except Exception:
                df["DATA ATT"] = df["DATA ATT"].fillna(pd.to_datetime("today").strftime("%Y-%m-%d"))
        else:
            df["DATA ATT"] = pd.to_datetime("today").strftime("%Y-%m-%d")

        # remover linhas totalmente vazias
        df = df.dropna(how="all").reset_index(drop=True)
        return df

    def read_sheet(self, sheet_name: str) -> pd.DataFrame:
        """
        Lê a sheet especificada:
        - detecta header (linha com mais células não-nulas entre 0..4)
        - aplica normalização
        - atualiza cache
        """
        try:
            # cache
            if not self._should_reload() and self._cache_df is not None:
                return self._cache_df.copy()

            if not os.path.exists(self.file_path):
                return pd.DataFrame()

            # ler sem header para detectar
            try:
                raw = pd.read_excel(self.file_path, sheet_name=sheet_name, engine="openpyxl", header=None, dtype=object)
            except Exception:
                # se sheet não existir ou erro, retornar df vazio
                return pd.DataFrame()

            header_row = 0
            best_non_null = -1
            max_check = min(5, len(raw))
            for i in range(max_check):
                non_null = raw.iloc[i].count()
                if non_null > best_non_null:
                    best_non_null = non_null
                    header_row = i

            # re-ler com header detectado
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine="openpyxl", header=header_row, dtype=object)

            # normalizar
            df = self._normalize_columns(df)

            # garantir colunas essenciais
            if "RENTABILIDADE_ULT_MES" not in df.columns:
                df["RENTABILIDADE_ULT_MES"] = 0.0

            # atualizar cache
            self._cache_df = df.copy()
            self._cache_mtime = self._file_mtime()
            self._last_cache_time = time.time()

            # antes de retornar df
            df = df.where(pd.notnull(df), None)
            return df

        except Exception as e:
            print(f"Erro inesperado ao ler a planilha '{sheet_name}': {e}")
            return pd.DataFrame()

    def write_sheet(self, df: pd.DataFrame, sheet_name: str):
        """
        Escreve DataFrame em uma aba (substitui se já existir).
        Usa openpyxl para compatibilidade.
        """
        try:
            if not os.path.exists(self.file_path):
                # criar workbook novo
                wb = load_workbook()
            else:
                wb = load_workbook(self.file_path)

            if sheet_name in wb.sheetnames:
                std = wb[sheet_name]
                wb.remove(std)
            ws = wb.create_sheet(title=sheet_name)

            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # ajustar largura simples
            for i, col in enumerate(df.columns, start=1):
                try:
                    ws.column_dimensions[get_column_letter(i)].auto_size = True
                except Exception:
                    pass

            wb.save(self.file_path)
        except Exception as e:
            print(f"Erro ao escrever a planilha '{sheet_name}': {e}")

    def update_sheet_range(self, sheet_name: str, start_row: int, start_col: int, df: pd.DataFrame):
        try:
            if not os.path.exists(self.file_path):
                self.write_sheet(df, sheet_name)
                return
            wb = load_workbook(self.file_path)
            if sheet_name not in wb.sheetnames:
                self.write_sheet(df, sheet_name)
                return
            ws = wb[sheet_name]
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
                for c_idx, value in enumerate(row, start=start_col):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            wb.save(self.file_path)
        except Exception as e:
            print(f"Erro ao atualizar o intervalo da planilha '{sheet_name}': {e}")

    def create_historical_sheet(self, df: pd.DataFrame):
        try:
            today_str = pd.to_datetime("today").strftime("%Y%m%d")
            sheet_name = f"base{today_str}"
            self.write_sheet(df, sheet_name)
            print(f"Aba histórica '{sheet_name}' criada com sucesso.")
        except Exception as e:
            print(f"Erro ao criar aba histórica: {e}")


